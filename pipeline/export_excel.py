"""
export_excel.py — Export validation results to a formatted Excel workbook.

Usage:
  python3 export_excel.py                                   # reads results/extracted.csv
  python3 export_excel.py --in results/extracted.csv --out results/validation.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── Colors ────────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_TP          = "C6EFCE"   # green
C_FP          = "FFCCCC"   # red
C_TN          = "BDD7EE"   # blue
C_FN          = "FFE699"   # yellow
C_EXCL        = "D9D9D9"   # grey
C_CANNOT      = "F2F2F2"   # light grey
C_UNRELIABLE  = "FFF2CC"   # amber
C_ACCENT      = "2E75B6"   # blue accent
C_SECTION_BG  = "D6E4F0"   # light blue section header

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _classify(row: dict) -> str:
    """Return TP/FP/TN/FN/EXCL/CANNOT/UNKNOWN."""
    if str(row.get("excluded", "")).lower() == "true":
        return "EXCL"
    pred = row.get("predicted_gg2_pos", "")
    actual = row.get("path_gg2_positive", "")
    if pred in ("", "None") or actual in ("", "None"):
        return "CANNOT"
    pred_b   = str(pred).lower() == "true"
    actual_b = str(actual).lower() == "true"
    if pred_b and actual_b:   return "TP"
    if pred_b and not actual_b: return "FP"
    if not pred_b and actual_b: return "FN"
    return "TN"


def _gg_label(row: dict) -> str:
    gg = row.get("path_gg_max", "")
    benign = str(row.get("path_benign", "")).lower() == "true"
    if gg not in ("", "None", None):
        try:
            return f"GG{int(float(gg))}"
        except Exception:
            pass
    if benign:
        return "Benign"
    return "—"


def _build_qa_sheet(wb, rows):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Dr. Tewari — Q&A")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    col_w = {"A":2,"B":3,"C":44,"D":2,"E":44,"F":2}
    for c,w in col_w.items():
        ws.column_dimensions[c].width = w

    def row_h(r,h): ws.row_dimensions[r].height = h
    def merge(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

    def s(r,c,v="",bold=False,sz=10,fg="222222",bg=None,ha="left",italic=False,wrap=True):
        x = ws.cell(row=r,column=c,value=v)
        x.font = Font(name="Arial",bold=bold,size=sz,color=fg,italic=italic)
        x.alignment = Alignment(horizontal="center" if ha=="center" else "left",
                                 vertical="top",wrap_text=wrap)
        if bg: x.fill = PatternFill("solid",fgColor=bg)
        return x

    # ── Pre-compute numbers ───────────────────────────────────────────────────
    pts = [r for r in rows
           if r.get('excluded','').lower() != 'true'
           and r.get('path_gg2_positive','') not in ('','None')]
    N = len(pts)
    excl = sum(1 for r in rows if r.get('excluded','').lower() == 'true')
    no_path = sum(1 for r in rows
                  if r.get('excluded','').lower() != 'true'
                  and r.get('path_gg2_positive','') in ('','None'))

    def psa(r): v=r.get('psa',''); return float(v) if v not in ('','None') else None
    def pirads(r): v=r.get('pirads',''); return int(v) if v not in ('','None') else None
    def psad(r): v=r.get('psad',''); return float(v) if v not in ('','None') else None
    def is_gg2(r): return r.get('path_gg2_positive','').lower()=='true'
    def is_hg(r):
        gg=r.get('path_gg_max',''); return int(float(gg))>=3 if gg not in ('','None') else False

    total_gg2 = sum(1 for r in pts if is_gg2(r))
    total_hg  = sum(1 for r in pts if is_hg(r))
    total_low = total_gg2 - total_hg  # GG2 only

    # PSA >= 2.4
    psa_ref   = [psa(r) is not None and psa(r)>=2.4 for r in pts]
    psa_bx    = sum(psa_ref)
    psa_c2    = sum(1 for r,b in zip(pts,psa_ref) if b and is_gg2(r))
    psa_m2    = sum(1 for r,b in zip(pts,psa_ref) if not b and is_gg2(r))
    psa_mhg   = sum(1 for r,b in zip(pts,psa_ref) if not b and is_hg(r))
    psa_un    = sum(1 for r,b in zip(pts,psa_ref) if b and not is_gg2(r))

    # MRI alone PI-RADS 4-5
    mri_ref   = [pirads(r) in (4,5) for r in pts]
    mri_bx    = sum(mri_ref)
    mri_c2    = sum(1 for r,b in zip(pts,mri_ref) if b and is_gg2(r))
    mri_m2    = sum(1 for r,b in zip(pts,mri_ref) if not b and is_gg2(r))
    mri_mhg   = sum(1 for r,b in zip(pts,mri_ref) if not b and is_hg(r))
    mri_un    = sum(1 for r,b in zip(pts,mri_ref) if b and not is_gg2(r))

    # PSA + MRI
    pm_ref    = [psa(r) is not None and psa(r)>=2.4 and pirads(r) in (4,5) for r in pts]
    pm_bx     = sum(pm_ref)
    pm_c2     = sum(1 for r,b in zip(pts,pm_ref) if b and is_gg2(r))
    pm_m2     = sum(1 for r,b in zip(pts,pm_ref) if not b and is_gg2(r))
    pm_mhg    = sum(1 for r,b in zip(pts,pm_ref) if not b and is_hg(r))
    pm_un     = sum(1 for r,b in zip(pts,pm_ref) if b and not is_gg2(r))

    # Best: PSA + (PI-RADS 4-5 OR PSAD>=0.15)
    best_ref  = [psa(r) is not None and psa(r)>=2.4 and
                 (pirads(r) in (4,5) or (psad(r) is not None and psad(r)>=0.15)) for r in pts]
    best_bx   = sum(best_ref)
    best_c2   = sum(1 for r,b in zip(pts,best_ref) if b and is_gg2(r))
    best_m2   = sum(1 for r,b in zip(pts,best_ref) if not b and is_gg2(r))
    best_mhg  = sum(1 for r,b in zip(pts,best_ref) if not b and is_hg(r))
    best_un   = sum(1 for r,b in zip(pts,best_ref) if b and not is_gg2(r))

    # ePSA model
    mdl_ref   = [str(r.get('predicted_gg2_pos','')).lower()=='true' for r in pts]
    mdl_bx    = sum(mdl_ref)
    mdl_c2    = sum(1 for r,b in zip(pts,mdl_ref) if b and is_gg2(r))
    mdl_m2    = sum(1 for r,b in zip(pts,mdl_ref) if not b and is_gg2(r))
    mdl_mhg   = sum(1 for r,b in zip(pts,mdl_ref) if not b and is_hg(r))
    mdl_un    = sum(1 for r,b in zip(pts,mdl_ref) if b and not is_gg2(r))

    # low risk (GG1) sent to biopsy
    lr_bx_all = sum(1 for r in pts if r.get('path_gg_max','') in ('0','1'))
    lr_bx_psa = sum(1 for r,b in zip(pts,psa_ref) if b and r.get('path_gg_max','') in ('0','1'))
    lr_bx_pm  = sum(1 for r,b in zip(pts,pm_ref)  if b and r.get('path_gg_max','') in ('0','1'))

    # ── Page layout ───────────────────────────────────────────────────────────
    row_h(1,6); row_h(2,40); row_h(3,16); row_h(4,6)
    merge(2,2,2,5)
    s(2,2,"Questions for Dr. Tewari — How Are Our Models Doing?",
      bold=True,sz=16,fg="1F3864")
    merge(3,2,3,5)
    s(3,2,f"Based on {len(rows)} patients processed  ·  {N} with confirmed biopsy outcomes  ·  {excl} excluded (post-prostatectomy)  ·  {no_path} missing pathology",
      sz=9,fg="888888",italic=True)

    # divider
    row_h(5,3); merge(5,2,5,5)
    ws.cell(row=5,column=2).fill = PatternFill("solid",fgColor="1F3864")

    QA = [
        # (Q#, question_text, answer_text, stat_line, bg_accent)
        ("Q1",
         "How good is PSA as a screening tool for our patients?",
         f"PSA alone (≥2.4 ng/mL) is a reasonable first filter. Of {N} biopsied patients, it would have referred {psa_bx} ({psa_bx/N*100:.0f}%) for biopsy. It caught {psa_c2} of {total_gg2} significant cancers ({psa_c2/total_gg2*100:.0f}% sensitivity) and only missed {psa_m2} — including {psa_mhg} high-grade (GG3+) cases.\n\nThe downside: {psa_un} of those {psa_bx} biopsies ({psa_un/psa_bx*100:.0f}%) came back negative for significant cancer — meaning most people sent to biopsy by PSA alone didn't need one. PSA is a good alarm bell but a poor decision-maker on its own.",
         f"Sensitivity {psa_c2/total_gg2*100:.0f}%  ·  {psa_m2} GG≥2 missed  ·  {psa_mhg} high-grade missed  ·  {psa_un/psa_bx*100:.0f}% unnecessary biopsies",
         "E3F2FD"),

        ("Q2",
         "How is MRI performing as a biopsy decision tool?",
         f"MRI (PI-RADS 4–5) sent {mri_bx} of {N} patients ({mri_bx/N*100:.0f}%) to biopsy and caught {mri_c2} of {total_gg2} significant cancers ({mri_c2/total_gg2*100:.0f}% sensitivity). It missed {mri_m2} GG≥2 cancers — {mri_mhg} of which were high-grade (GG3+).\n\nThe missed cases are important: these are patients with confirmed significant cancer whose MRI did not show a PI-RADS 4–5 lesion. They were likely found by systematic or PSMA-guided biopsy. This is a known limitation of MRI-only triage — it has a blind spot for cancers without a discrete focal lesion.\n\nOf those sent to biopsy by MRI, {mri_un} ({mri_un/mri_bx*100:.0f}%) had no significant cancer — so MRI still generates substantial unnecessary biopsies.",
         f"Sensitivity {mri_c2/total_gg2*100:.0f}%  ·  {mri_m2} GG≥2 missed  ·  {mri_mhg} high-grade missed  ·  {mri_un/mri_bx*100:.0f}% unnecessary biopsies",
         "FFF8E1"),

        ("Q3",
         "What happens when we combine PSA + MRI together as the biopsy gate?",
         f"Requiring both PSA ≥2.4 AND PI-RADS 4–5 cuts biopsy referrals significantly — only {pm_bx} of {N} patients ({pm_bx/N*100:.0f}%) would have been sent. But the trade-off is real: {pm_m2} significant cancers are missed, including {pm_mhg} high-grade cases.\n\nThese missed high-grade cancers are the concern. They include patients whose PSA was elevated but MRI was PI-RADS 1–3, or whose MRI was positive but PSA was below threshold. Adding PSAD ≥0.15 as an alternative trigger rescues most of these — see Q5.",
         f"Sensitivity {pm_c2/total_gg2*100:.0f}%  ·  {pm_m2} GG≥2 missed  ·  {pm_mhg} high-grade missed  ·  saves {N-pm_bx-N+psa_bx} more biopsies vs PSA alone",
         "F3E5F5"),

        ("Q4",
         "Low-risk vs high-risk biopsies — are we sending the right people?",
         f"Of the {N} patients who actually got biopsied:\n  •  {total_gg2} ({total_gg2/N*100:.0f}%) had clinically significant cancer (GG≥2) — these were necessary\n  •  {total_hg} ({total_hg/N*100:.0f}%) had high-grade cancer (GG3+) — most urgently needed biopsy\n  •  {total_low} ({total_low/N*100:.0f}%) had GG2 only (3+4) — significant but lower urgency\n  •  {lr_bx_all} ({lr_bx_all/N*100:.0f}%) had GG0 or GG1 — benign or indolent cancer, biopsy arguably avoidable\n\nIf we had used PSA ≥2.4 + PI-RADS 4–5 as the gate, we would have still sent {lr_bx_pm} low-risk patients to biopsy unnecessarily. The model does not yet reliably distinguish GG1 from GG2.",
         f"{total_hg} high-grade  ·  {total_low} GG2  ·  {lr_bx_all} low-risk/benign biopsied  ·  {lr_bx_all/N*100:.0f}% potentially avoidable",
         "E8F5E9"),

        ("Q5",
         "How many bad cancers did we miss? And with which strategy?",
         f"High-grade cancer (GG3+) is the one we cannot afford to miss. There were {total_hg} GG3+ cases in this cohort.\n\n  •  Biopsy ALL: 0 missed (gold standard, 100%)\n  •  PSA ≥2.4 only: {psa_mhg} missed\n  •  MRI alone (PI-RADS 4–5): {mri_mhg} missed\n  •  PSA + MRI combined: {pm_mhg} missed\n  •  PSA + (MRI OR PSAD ≥0.15): {best_mhg} missed ← recommended\n  •  ePSA model: {mdl_mhg} missed\n\nThe recommended pathway (PSA + MRI-or-PSAD) misses only {best_mhg} high-grade cancer(s) while avoiding {N-best_bx} biopsies. That is the best trade-off in this dataset.",
         f"Worst: MRI/PSA+MRI miss {mri_mhg}/{pm_mhg} high-grade  ·  Best trade-off: PSA+(MRI or PSAD) misses {best_mhg}",
         "FFEBEE"),

        ("Q6",
         "Is the ePSA model doing the right thing?",
         f"The model was trained on a high-risk cohort (74% GG≥2 prevalence). In this real-world validation cohort the true GG≥2 rate is only {total_gg2/N*100:.0f}% — so the model is calibrated for a much sicker population than it's now being tested on.\n\nResult: it predicts GG≥2 for nearly everyone with a PI-RADS score, regardless of PSA or PSAD. It referred {mdl_bx} of {N} patients ({mdl_bx/N*100:.0f}%) to biopsy, caught {mdl_c2} of {total_gg2} cancers ({mdl_c2/total_gg2*100:.0f}%), but generated {mdl_un} false positives ({mdl_un/mdl_bx*100:.0f}% of those referred).\n\nThe model's best use right now is risk communication — showing a patient their probability number — not as a binary biopsy gate. It needs retraining on a cohort that matches the prevalence of our actual patient population (~28% GG≥2).",
         f"Sensitivity {mdl_c2/total_gg2*100:.0f}%  ·  {mdl_m2} GG≥2 missed  ·  {mdl_mhg} high-grade missed  ·  {mdl_un/mdl_bx*100:.0f}% false positive rate",
         "E8EAF6"),

        ("Q7",
         "Bottom line: which approach should we use?",
         f"Based on this {N}-patient dataset, the recommended clinical pathway is:\n\n  PSA ≥ 2.4  AND  (PI-RADS 4–5  OR  PSAD ≥ 0.15)\n\nThis catches {best_c2} of {total_gg2} significant cancers ({best_c2/total_gg2*100:.0f}% sensitivity), misses only {best_mhg} high-grade case(s), and avoids {N-best_bx} ({(N-best_bx)/N*100:.0f}%) unnecessary biopsies.\n\nThe ePSA model, once retrained on a representative cohort, could sharpen this further — particularly for borderline PSAD cases. But with the current training data, PSA + MRI + PSAD together outperforms the model as a biopsy decision tool.",
         f"Recommended: PSA ≥2.4 + (PI-RADS 4–5 OR PSAD ≥0.15)  ·  {best_c2}/{total_gg2} caught  ·  {best_mhg} high-grade missed  ·  {(N-best_bx)/N*100:.0f}% biopsies avoided",
         "E8F5E9"),
    ]

    row_cur = 7
    for num, question, answer, stat, bg in QA:
        # Question bar
        row_h(row_cur, 4)
        row_cur += 1

        row_h(row_cur, 22)
        merge(row_cur, 2, row_cur, 5)
        x = ws.cell(row=row_cur, column=2, value=f"  {num}   {question}")
        x.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        x.fill = PatternFill("solid", fgColor="1F3864")
        x.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        row_cur += 1

        # Answer block — split left/right
        answer_lines = answer.split('\n')
        n_lines = max(len(answer_lines), 2)
        h = max(90, n_lines * 14)
        row_h(row_cur, h)

        merge(row_cur, 2, row_cur, 3)
        x = ws.cell(row=row_cur, column=2, value=answer)
        x.font = Font(name="Arial", size=10, color="333333")
        x.fill = PatternFill("solid", fgColor=bg)
        x.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)

        # Stat pill on right
        merge(row_cur, 5, row_cur, 5)
        x2 = ws.cell(row=row_cur, column=5, value=f"📊  {stat}")
        x2.font = Font(name="Arial", size=9, color="1F3864", bold=True)
        x2.fill = PatternFill("solid", fgColor="DDEEFF")
        x2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)

        row_cur += 1

    # Footer
    row_h(row_cur, 4); row_cur += 1
    merge(row_cur, 2, row_cur, 5)
    ws.cell(row=row_cur, column=2).fill = PatternFill("solid", fgColor="1F3864")
    row_h(row_cur, 3); row_cur += 1
    row_h(row_cur, 14)
    merge(row_cur, 2, row_cur, 5)
    x = ws.cell(row=row_cur, column=2,
                 value="ePSA Model Validation  ·  Mount Sinai Urology  ·  Confidential — for internal use only")
    x.font = Font(name="Arial", size=8, color="AAAAAA", italic=True)
    x.alignment = Alignment(horizontal="center", vertical="center")


def _build_pathway_sheet(wb, rows):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Pathway Analysis")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # Column widths
    col_w = {"A":2,"B":6,"C":32,"D":12,"E":12,"F":12,"G":12,"H":12,"I":12,"J":2,"K":2,"L":36}
    for c,w in col_w.items():
        ws.column_dimensions[c].width = w

    def s(r,c,v="",bold=False,sz=10,fg="222222",bg=None,ha="left",italic=False,wrap=True,border=False):
        x = ws.cell(row=r,column=c,value=v)
        x.font = Font(name="Arial",bold=bold,size=sz,color=fg,italic=italic)
        x.alignment = Alignment(horizontal="center" if ha=="center" else ("right" if ha=="right" else "left"),
                                 vertical="center",wrap_text=wrap)
        if bg: x.fill = PatternFill("solid",fgColor=bg)
        if border:
            t = Side(style="thin",color="CCCCCC")
            x.border = Border(left=t,right=t,top=t,bottom=t)
        return x

    def row_h(r,h): ws.row_dimensions[r].height = h
    def merge(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

    # ── Title ────────────────────────────────────────────────────────────────
    row_h(1,6); row_h(2,36); row_h(3,18); row_h(4,6)
    merge(2,2,2,12)
    s(2,2,"Biopsy Pathway Comparison — What Would Each Strategy Have Caught?",
      bold=True,sz=15,fg="1F3864",ha="left")
    merge(3,2,3,12)
    s(3,2,"149 patients with confirmed biopsy outcomes  ·  42 GG≥2 (28%)  ·  19 GG3+ high-grade (13%)  ·  Mount Sinai Urology",
      sz=9,fg="888888",italic=True)

    # ── Column headers ───────────────────────────────────────────────────────
    row_h(5,4)
    row_h(6,32)
    hdrs = ["#","Pathway / Strategy","Biopsied\n(of 149)","GG≥2\nCaught","GG≥2\nMissed","HG (GG3+)\nMissed","Unnecessary\nBiopsies","PPV\n(hit rate)"]
    hcols = [2,3,4,5,6,7,8,9]
    for ci,h in zip(hcols,hdrs):
        x = s(6,ci,h,bold=True,sz=9,fg="FFFFFF",bg="1F3864",ha="center",wrap=True)

    # ── Pathway data ─────────────────────────────────────────────────────────
    pts = [r for r in rows
           if r.get('excluded','').lower() != 'true'
           and r.get('path_gg2_positive','') not in ('','None')]
    N = len(pts)

    def psa(r):
        v=r.get('psa',''); return float(v) if v not in ('','None') else None
    def pirads(r):
        v=r.get('pirads',''); return int(v) if v not in ('','None') else None
    def psad(r):
        v=r.get('psad',''); return float(v) if v not in ('','None') else None
    def is_gg2(r): return r.get('path_gg2_positive','').lower()=='true'
    def is_hg(r):
        gg=r.get('path_gg_max',''); return int(float(gg))>=3 if gg not in ('','None') else False

    total_gg2 = sum(1 for r in pts if is_gg2(r))
    total_hg  = sum(1 for r in pts if is_hg(r))

    def calc(referred):
        bx  = sum(referred)
        c2  = sum(1 for r,b in zip(pts,referred) if b and is_gg2(r))
        m2  = sum(1 for r,b in zip(pts,referred) if not b and is_gg2(r))
        mhg = sum(1 for r,b in zip(pts,referred) if not b and is_hg(r))
        un  = sum(1 for r,b in zip(pts,referred) if b and not is_gg2(r))
        ppv = c2/bx if bx else 0
        sens= c2/total_gg2 if total_gg2 else 0
        return bx, c2, m2, mhg, un, ppv, sens

    pathways = [
        ("1", "Biopsy ALL  (status quo)\nEveryone referred regardless of PSA, MRI, or risk.",
         [True]*N, "FFFFFF", False),
        ("2", "PSA ≥ 2.4 ng/mL  (PSA screen only)\nOnly patients with PSA above the threshold are sent for biopsy. No MRI required.",
         [psa(r) is not None and psa(r)>=2.4 for r in pts], "FFFFFF", False),
        ("3", "MRI alone  (PI-RADS 4–5)\nBiopsy only if mpMRI shows a suspicious lesion (PI-RADS 4 or 5). No PSA gate.",
         [pirads(r) in (4,5) for r in pts], "FFFFFF", False),
        ("4", "PSA ≥ 2.4  +  PI-RADS 4–5\nTwo-gate: elevated PSA AND suspicious MRI both required before biopsy.",
         [psa(r) is not None and psa(r)>=2.4 and pirads(r) in (4,5) for r in pts], "EBF3FB", False),
        ("5", "PSA ≥ 2.4  +  PSAD ≥ 0.10\nPSA density added as second filter. PSAD <0.10 has 94% NPV for ruling out GG≥2.",
         [psa(r) is not None and psa(r)>=2.4 and psad(r) is not None and psad(r)>=0.10 for r in pts], "FFFFFF", False),
        ("6", "PSA ≥ 2.4  +  (PI-RADS 4–5  OR  PSAD ≥ 0.15)  ★ Best balance\nBroad safety net: biopsy if MRI is suspicious OR density is high. Either flag triggers referral.",
         [psa(r) is not None and psa(r)>=2.4 and
          (pirads(r) in (4,5) or (psad(r) is not None and psad(r)>=0.15)) for r in pts], "E8F5E9", True),
        ("7", "ePSA Model  (P(GG≥2) ≥ 50%)\nLogistic regression using PSA + PI-RADS dummy variables predicts risk of clinically significant cancer.",
         [str(r.get('predicted_gg2_pos','')).lower()=='true' for r in pts], "FFFFFF", False),
        ("8", "PSA ≥ 2.4  +  PI-RADS 4–5  +  PSAD ≥ 0.10  (triple gate)\nAll three criteria must be met. Most restrictive — maximises biopsy avoidance but risks missing cancers.",
         [psa(r) is not None and psa(r)>=2.4 and pirads(r) in (4,5) and
          psad(r) is not None and psad(r)>=0.10 for r in pts], "FFFFFF", False),
    ]

    for i,(num,desc,ref,row_bg,highlight) in enumerate(pathways):
        ri = 7 + i*2
        row_h(ri, 42); row_h(ri+1, 4)
        bx,c2,m2,mhg,un,ppv,sens = calc(ref)
        bg = "FFFDE7" if highlight else row_bg
        border_col = "F9A825" if highlight else None

        s(ri,2,num,bold=True,sz=10,fg="1F3864",bg=bg,ha="center",wrap=False)
        s(ri,3,desc.split('\n')[0],bold=True,sz=10,fg="1F3864" if not highlight else "2E7D32",bg=bg,wrap=False)

        vals = [
            f"{bx}  ({bx/N*100:.0f}%)",
            f"{c2}/{total_gg2}  ({sens*100:.0f}%)",
            str(m2),
            str(mhg),
            f"{un}  ({un/bx*100:.0f}%)" if bx else "—",
            f"{ppv*100:.0f}%",
        ]
        v_cols   = [4,5,6,7,8,9]
        v_fgs    = ["333333",
                    "1A6B1A" if sens>=0.85 else ("E07F00" if sens>=0.75 else "9C0006"),
                    "9C0006" if m2>5 else ("E07F00" if m2>2 else "1A6B1A"),
                    "9C0006" if mhg>5 else ("E07F00" if mhg>2 else "1A6B1A"),
                    "333333",
                    "1A6B1A" if ppv>=0.4 else "333333"]
        for ci2,v,vfg in zip(v_cols,vals,v_fgs):
            s(ri,ci2,v,bold=(ci2 in (5,6)),sz=10,fg=vfg,bg=bg,ha="center",border=True,wrap=False)

        # Explanation text in column L
        explanation = desc.split('\n')[1] if '\n' in desc else ""
        merge(ri,12,ri,12)
        s(ri,12,explanation,sz=9,fg="555555",bg="FAFAFA",italic=True,wrap=True)

    # ── Key insights box ──────────────────────────────────────────────────────
    start = 7 + len(pathways)*2 + 2
    row_h(start,4)
    row_h(start+1,24)
    merge(start+1,2,start+1,12)
    s(start+1,2,"KEY INSIGHTS",bold=True,sz=9,fg="FFFFFF",bg="1F3864",ha="left")

    insights = [
        ("Missed high-grade cancers (GG3+) are the critical metric.",
         "Missing a GG3 or above is a serious harm. Any pathway that misses >5 high-grade cancers needs careful justification."),
        ("PSA alone (≥2.4) is surprisingly effective.",
         "Catches 93% of GG≥2 with only 3 missed — but still sends 87% of patients to biopsy with a 30% hit rate."),
        ("MRI (PI-RADS 4–5) alone misses 8 GG≥2 including 7 high-grade.",
         "These are likely PSMA-detected or systematic biopsy cases without a discrete MRI lesion — a real limitation."),
        ("★ Pathway 6 is the best balance.",
         "PSA ≥ 2.4 + (PI-RADS 4–5 OR PSAD ≥ 0.15) catches 88% of GG≥2, avoids 32% of biopsies, and misses only 4 high-grade cancers."),
        ("Triple gate (Pathway 8) is too restrictive.",
         "Misses 9 of 19 high-grade cancers (47%). The biopsy reduction (59%) does not justify this miss rate."),
        ("The ePSA model adds limited value as a standalone tool.",
         "83% sensitivity, but refers 81% of patients — barely better than PSA alone. Most useful as a risk communication tool, not a gating decision."),
    ]

    for j,(title,body) in enumerate(insights):
        ri2 = start+2+j
        row_h(ri2, 28)
        merge(ri2,2,ri2,2)
        s(ri2,2,f"{j+1}",bold=True,sz=10,fg="FFFFFF",bg="2E75B6",ha="center",wrap=False)
        merge(ri2,3,ri2,6)
        s(ri2,3,title,bold=True,sz=10,fg="1F3864",bg="F7F9FC",wrap=False)
        merge(ri2,7,ri2,12)
        s(ri2,7,body,sz=9,fg="444444",bg="F7F9FC",wrap=True)

    # ── Footer divider ────────────────────────────────────────────────────────
    last = start+2+len(insights)+1
    row_h(last,3); merge(last,2,last,12)
    ws.cell(row=last,column=2).fill = PatternFill("solid",fgColor="1F3864")
    row_h(last+1,14); merge(last+1,2,last+1,12)
    x=ws.cell(row=last+1,column=2,
               value="ePSA Pathway Analysis  ·  Mount Sinai Urology  ·  Confidential — for internal use only")
    x.font=Font(name="Arial",size=8,color="AAAAAA",italic=True)
    x.alignment=Alignment(horizontal="center",vertical="center")


def _build_tewari_sheet(wb, rows, counts, total, excl, evaluated,
                        tp, fp, tn, fn, precision, sensitivity, specificity,
                        reliable_total, reliable_correct,
                        unreliable_total, unreliable_correct):
    ws = wb.create_sheet("Model Validation — Summary")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # Column widths
    col_widths = {"A":3,"B":28,"C":14,"D":14,"E":14,"F":14,"G":3}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Row heights (preset for clean layout)
    for r in range(1, 55):
        ws.row_dimensions[r].height = 18
    for r in [1, 2, 4, 14, 24, 34, 44]:
        ws.row_dimensions[r].height = 8   # spacer rows

    def cell(r, c, val="", bold=False, size=11, color="222222", bg=None,
             align="left", border=False, italic=False, wrap=False):
        x = ws.cell(row=r, column=c, value=val)
        x.font = Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
        x.alignment = Alignment(
            horizontal="center" if align == "center" else ("right" if align == "right" else "left"),
            vertical="center", wrap_text=wrap
        )
        if bg:
            x.fill = PatternFill("solid", fgColor=bg)
        if border:
            thin = Side(style="thin", color="CCCCCC")
            x.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return x

    def section_header(r, label):
        ws.merge_cells(f"B{r}:F{r}")
        c = cell(r, 2, label.upper(), bold=True, size=9, color="FFFFFF", bg="1F3864", align="left")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 22

    def metric_row(r, label, value, note="", value_color="222222", bg="FFFFFF"):
        cell(r, 2, label, size=10, color="444444", bg=bg)
        cell(r, 3, value, bold=True, size=11, color=value_color, bg=bg, align="center")
        if note:
            x = ws.cell(row=r, column=4)
            ws.merge_cells(f"D{r}:F{r}")
            cell(r, 4, note, size=9, color="888888", bg=bg, italic=True)

    # ── Title block ──────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 40
    ws.merge_cells("B3:F3")
    cell(3, 2, "GG≥2 Prostate Cancer Detection — Model Validation",
         bold=True, size=16, color="1F3864", align="left")

    ws.row_dimensions[5].height = 14
    ws.merge_cells("B5:F5")
    cell(5, 2, "Mount Sinai Urology  ·  Biopsy Registry  ·  Logistic Regression Model (N = 96 training cohort)",
         size=10, color="888888", italic=True)

    ws.row_dimensions[6].height = 14
    ws.merge_cells("B6:F6")
    cell(6, 2, "Model: ln(PSA) + PI-RADS dummy variables  ·  AUC 0.591  ·  Outcome: GG≥2 (clinically significant PCa)",
         size=10, color="888888", italic=True)

    # divider
    ws.row_dimensions[7].height = 4
    ws.merge_cells("B8:F8")
    div = ws.cell(row=8, column=2)
    div.fill = PatternFill("solid", fgColor="1F3864")
    ws.row_dimensions[8].height = 3

    # ── Section 1: Cohort ────────────────────────────────────────────────────
    section_header(10, "Validation Cohort")
    cannot = counts["CANNOT"]
    cohort_rows = [
        ("Total patients processed",              total,      ""),
        ("Excluded (post-prostatectomy / other)", excl,       "S/P RALP, TURP, cystoscopy-only cases"),
        ("Missing PSA or PI-RADS (no prediction)",cannot,    "Cannot run model without both inputs"),
        ("Evaluated by model",                    evaluated,  ""),
    ]
    for i, (lbl, val, note) in enumerate(cohort_rows, 11):
        bg = "F7F9FC" if i % 2 == 0 else "FFFFFF"
        metric_row(i, lbl, val, note, bg=bg)

    # ── Section 2: Confusion Matrix ──────────────────────────────────────────
    section_header(16, "Confusion Matrix  (threshold: P ≥ 50%)")

    # Header row
    for c_idx, (col, label) in enumerate([(3,"Predicted GG≥2"),(4,"Predicted GG<2"),(5,"Total")], 3):
        x = cell(17, c_idx, label, bold=True, size=9, color="FFFFFF", bg="2E75B6", align="center")
    ws.row_dimensions[17].height = 22

    # Actual GG≥2 row
    cell(18, 2, "Actual GG≥2", bold=True, size=10, color="FFFFFF", bg="2E75B6", align="center")
    cell(18, 3, tp, bold=True, size=13, color="1A6B1A", bg="C6EFCE", align="center", border=True)
    cell(18, 4, fn, bold=True, size=13, color="7F4700", bg="FFE699", align="center", border=True)
    cell(18, 5, f"={tp+fn}", bold=True, size=11, color="333333", bg="EBF3FB", align="center")
    ws.row_dimensions[18].height = 32

    # Actual GG<2 row
    cell(19, 2, "Actual GG<2", bold=True, size=10, color="FFFFFF", bg="2E75B6", align="center")
    cell(19, 3, fp, bold=True, size=13, color="9C0006", bg="FFCCCC", align="center", border=True)
    cell(19, 4, tn, bold=True, size=13, color="1F497D", bg="BDD7EE", align="center", border=True)
    cell(19, 5, f"={fp+tn}", bold=True, size=11, color="333333", bg="EBF3FB", align="center")
    ws.row_dimensions[19].height = 32

    # Total row
    cell(20, 2, "Total", bold=True, size=10, color="333333", bg="EBF3FB")
    cell(20, 3, tp+fp, bold=True, size=11, color="333333", bg="EBF3FB", align="center")
    cell(20, 4, fn+tn, bold=True, size=11, color="333333", bg="EBF3FB", align="center")
    cell(20, 5, tp+fp+fn+tn, bold=True, size=11, color="333333", bg="EBF3FB", align="center")
    ws.row_dimensions[20].height = 22

    # TP/FP labels
    ws.merge_cells("C21:C21")
    cell(21, 3, "True Positive", size=8, color="1A6B1A", align="center", italic=True)
    cell(21, 4, "False Negative", size=8, color="7F4700", align="center", italic=True)
    ws.row_dimensions[21].height = 14
    cell(22, 3, "False Positive", size=8, color="9C0006", align="center", italic=True)
    cell(22, 4, "True Negative", size=8, color="1F497D", align="center", italic=True)
    ws.row_dimensions[22].height = 14

    # ── Section 3: Performance Metrics ───────────────────────────────────────
    section_header(24, "Overall Performance")

    overall_metrics = [
        ("Sensitivity (Recall)",
         f"{sensitivity:.0%}" if sensitivity is not None else "—",
         "Of all true GG≥2, % correctly predicted",
         "1A6B1A" if sensitivity and sensitivity >= 0.7 else "9C0006"),
        ("Specificity",
         f"{specificity:.0%}" if specificity is not None else "—",
         "Of all true GG<2, % correctly predicted",
         "1A6B1A" if specificity and specificity >= 0.5 else "9C0006"),
        ("Precision (PPV)",
         f"{precision:.0%}" if precision is not None else "—",
         "Of predicted GG≥2, % that were truly GG≥2",
         "1A6B1A" if precision and precision >= 0.6 else "9C0006"),
        ("AUC (training cohort)",
         "0.591",
         "Area under ROC curve — N=96 Mount Sinai registry",
         "E07F00"),
    ]
    for i, (lbl, val, note, vcol) in enumerate(overall_metrics, 25):
        bg = "F7F9FC" if i % 2 == 0 else "FFFFFF"
        metric_row(i, lbl, val, note, value_color=vcol, bg=bg)

    # ── Section 4: By PI-RADS Reliability ───────────────────────────────────
    section_header(30, "Performance by PI-RADS Reliability")

    rel_pct   = f"{reliable_correct/reliable_total:.0%}"   if reliable_total   else "—"
    unrel_pct = f"{unreliable_correct/unreliable_total:.0%}" if unreliable_total else "—"

    pirads_rows = [
        ("PI-RADS 4–5  (model reliable)",
         f"{reliable_correct}/{reliable_total}  ({rel_pct})",
         "Model trained on 74% GG≥2 prevalence — more accurate here",
         "1A6B1A"),
        ("PI-RADS ≤3  (model unreliable)",
         f"{unreliable_correct}/{unreliable_total}  ({unrel_pct})",
         "Overcalls GG≥2 at low PI-RADS — use PSAD and AUA rates instead",
         "9C0006"),
        ("Missing PI-RADS  (cannot predict)",
         str(counts["CANNOT"]),
         "No PI-RADS score extracted — excluded from model",
         "888888"),
    ]
    for i, (lbl, val, note, vcol) in enumerate(pirads_rows, 31):
        bg = "F7F9FC" if i % 2 == 0 else "FFFFFF"
        metric_row(i, lbl, val, note, value_color=vcol, bg=bg)

    # ── Section 5: Key Caveats ───────────────────────────────────────────────
    section_header(35, "Key Caveats & Limitations")

    caveats = [
        "Model calibrated to 74% GG≥2 prevalence in biopsied cohort — overcalls cancer at PI-RADS ≤3.",
        "AUC 0.591 is modest; intended as one input to shared decision-making, not standalone triage.",
        "N=96 training cohort (Mount Sinai) — external validation in progress.",
        "PSAD <0.10 ng/mL² (NPV 94%) should override model output at PI-RADS ≤3.",
        "Post-prostatectomy cases (S/P RALP, fossa biopsy) excluded from model — different biology.",
        "De-identification performed with OpenMed prior to extraction; names and dates suppressed.",
    ]
    for i, txt in enumerate(caveats, 36):
        ws.row_dimensions[i].height = 16
        ws.merge_cells(f"B{i}:F{i}")
        x = ws.cell(row=i, column=2, value=f"  •  {txt}")
        x.font = Font(name="Arial", size=9, color="444444")
        x.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        x.fill = PatternFill("solid", fgColor="F7F9FC" if i % 2 == 0 else "FFFFFF")

    # ── Footer ───────────────────────────────────────────────────────────────
    ws.row_dimensions[43].height = 3
    ws.merge_cells("B44:F44")
    foot = ws.cell(row=44, column=2)
    foot.fill = PatternFill("solid", fgColor="1F3864")
    ws.row_dimensions[44].height = 3

    ws.row_dimensions[45].height = 16
    ws.merge_cells("B45:F45")
    x = ws.cell(row=45, column=2,
                 value="ePSA Model Validation  ·  Mount Sinai Urology  ·  Confidential — for internal use only")
    x.font = Font(name="Arial", size=8, color="AAAAAA", italic=True)
    x.alignment = Alignment(horizontal="center", vertical="center")


def build_excel(in_csv: str, out_xlsx: str) -> None:
    rows = list(csv.DictReader(open(in_csv, encoding="utf-8")))

    wb = Workbook()

    # ── Sheet 1: Patient Results ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Patient Results"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    COLS = [
        ("Patient ID",      "patient_id",        9),
        ("Case Type",       "case_type",          28),
        ("PSA\n(ng/mL)",    "psa",                8),
        ("PI-RADS",         "pirads",             8),
        ("PSAD\n(ng/mL²)", "psad",               9),
        ("Vol\n(cc)",       "prostate_volume_cc", 8),
        ("Actual\nGG",      None,                 8),
        ("GG≥2\nActual",   "path_gg2_positive",  9),
        ("Cribriform",      "path_cribriform",    9),
        ("Predicted\nP(GG≥2)", "predicted_prob", 10),
        ("Predicted\nGG≥2","predicted_gg2_pos",  9),
        ("Model\nReliable", "model_reliable",     9),
        ("Result",          None,                 8),
        ("Notes",           "extraction_notes",   28),
    ]

    # Header row
    for ci, (label, _, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill      = _fill(C_HEADER_BG)
        cell.font      = _font(bold=True, color=C_HEADER_FG, size=10)
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 36

    # Data rows
    for ri, row in enumerate(rows, 2):
        cls = _classify(row)

        # Row background
        bg = {
            "TP":     C_TP,
            "FP":     C_FP,
            "TN":     C_TN,
            "FN":     C_FN,
            "EXCL":   C_EXCL,
            "CANNOT": C_CANNOT,
        }.get(cls, "FFFFFF")

        # Override with amber if unreliable
        if cls not in ("EXCL", "CANNOT") and str(row.get("model_reliable", "")).lower() == "false":
            bg = C_UNRELIABLE

        def _cell(col_idx: int, value, fmt=None, bold=False, align="center"):
            c = ws.cell(row=ri, column=col_idx, value=value)
            c.fill      = _fill(bg)
            c.font      = _font(bold=bold, size=10)
            c.alignment = _center() if align == "center" else _left()
            c.border    = _border()
            if fmt:
                c.number_format = fmt
            return c

        predicted_pct = None
        if row.get("predicted_prob") not in ("", "None", None):
            try:
                predicted_pct = float(row["predicted_prob"]) * 100
            except Exception:
                pass

        actual_gg2 = str(row.get("path_gg2_positive", "")).lower()
        crib       = str(row.get("path_cribriform", "")).lower()
        reliable   = str(row.get("model_reliable", "")).lower()

        values = [
            row.get("patient_id", ""),
            row.get("case_type", "")[:60],
            float(row["psa"]) if row.get("psa") not in ("", "None", None) else None,
            int(row["pirads"]) if row.get("pirads") not in ("", "None", None) else None,
            float(row["psad"]) if row.get("psad") not in ("", "None", None) else None,
            float(row["prostate_volume_cc"]) if row.get("prostate_volume_cc") not in ("", "None", None) else None,
            _gg_label(row),
            "Yes" if actual_gg2 == "true" else ("No" if actual_gg2 == "false" else "—"),
            "Yes" if crib == "true" else "No",
            predicted_pct,
            "Yes" if str(row.get("predicted_gg2_pos", "")).lower() == "true" else (
                "No" if str(row.get("predicted_gg2_pos", "")).lower() == "false" else "—"),
            "✓" if reliable == "true" else ("⚠" if reliable == "false" else "—"),
            cls,
            row.get("extraction_notes", ""),
        ]

        aligns = ["center","left","center","center","center","center",
                  "center","center","center","center","center","center","center","left"]
        fmts   = [None,None,"0.00",None,"0.000","0.0",None,None,None,"0.0\"%\"",None,None,None,None]

        for ci, (val, align, fmt) in enumerate(zip(values, aligns, fmts), 1):
            bold = ci == 13  # bold result column
            _cell(ci, val, fmt=fmt, bold=bold, align=align)

        # Color the Result cell separately
        result_cell = ws.cell(row=ri, column=13)
        result_colors = {"TP": ("006100","C6EFCE"), "FP": ("9C0006","FFCCCC"),
                         "TN": ("1F497D","BDD7EE"), "FN": ("7F6000","FFE699"),
                         "EXCL": ("595959","D9D9D9"), "CANNOT": ("595959","F2F2F2")}
        if cls in result_colors:
            fg, rbg = result_colors[cls]
            result_cell.fill = _fill(rbg)
            result_cell.font = _font(bold=True, color=fg, size=10)

        ws.row_dimensions[ri].height = 18

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 22

    def _s(r, c, val, bold=False, bg=None, align="left", size=11):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font      = _font(bold=bold, size=size)
        cell.alignment = _center() if align == "center" else _left()
        if bg:
            cell.fill = _fill(bg)
        cell.border = _border()
        return cell

    # Title
    ws2.merge_cells("A1:C1")
    t = ws2.cell(row=1, column=1, value="GG≥2 Model Validation — Summary")
    t.fill = _fill(C_HEADER_BG)
    t.font = _font(bold=True, color="FFFFFF", size=14)
    t.alignment = _center()
    ws2.row_dimensions[1].height = 30

    counts = {k: 0 for k in ["TP","FP","TN","FN","EXCL","CANNOT"]}
    reliable_total = reliable_correct = 0
    unreliable_total = unreliable_correct = 0

    for row in rows:
        cls = _classify(row)
        counts[cls] += 1
        rel = str(row.get("model_reliable","")).lower()
        correct = str(row.get("correct","")).lower()
        if cls not in ("EXCL","CANNOT"):
            if rel == "true":
                reliable_total += 1
                if correct == "true": reliable_correct += 1
            elif rel == "false":
                unreliable_total += 1
                if correct == "true": unreliable_correct += 1

    total    = len(rows)
    excl     = counts["EXCL"]
    evaluated= total - excl - counts["CANNOT"]
    tp, fp, tn, fn = counts["TP"], counts["FP"], counts["TN"], counts["FN"]
    precision   = tp/(tp+fp)   if (tp+fp) > 0  else None
    sensitivity = tp/(tp+fn)   if (tp+fn) > 0  else None
    specificity = tn/(tn+fp)   if (tn+fp) > 0  else None

    data = [
        ("── Cohort", None, None, C_SECTION_BG),
        ("Total patients",                total,    None, None),
        ("Excluded (post-proc / other)",  excl,     None, None),
        ("Cannot predict (missing data)", counts["CANNOT"], None, None),
        ("Evaluated",                     evaluated, None, None),
        ("Reliable predictions (PI-RADS 4–5)", reliable_total, None, None),
        ("Unreliable predictions (PI-RADS ≤3)", unreliable_total, None, None),
        ("── Confusion Matrix", None, None, C_SECTION_BG),
        ("True Positive (TP)",   tp,  "Predicted GG≥2 & actual GG≥2", C_TP),
        ("False Positive (FP)",  fp,  "Predicted GG≥2 & actual GG<2",  C_FP),
        ("True Negative (TN)",   tn,  "Predicted GG<2 & actual GG<2",  C_TN),
        ("False Negative (FN)",  fn,  "Predicted GG<2 & actual GG≥2",  C_FN),
        ("── Performance", None, None, C_SECTION_BG),
        ("Precision",    f"{precision:.2f}"   if precision   is not None else "—", None, None),
        ("Sensitivity",  f"{sensitivity:.2f}" if sensitivity is not None else "—", None, None),
        ("Specificity",  f"{specificity:.2f}" if specificity is not None else "—", None, None),
        ("── By Reliability", None, None, C_SECTION_BG),
        ("Reliable accuracy",   f"{reliable_correct}/{reliable_total} ({100*reliable_correct//reliable_total if reliable_total else 0}%)", None, None),
        ("Unreliable accuracy", f"{unreliable_correct}/{unreliable_total} ({100*unreliable_correct//unreliable_total if unreliable_total else 0}%)", None, None),
    ]

    for ri2, (label, val, note, bg) in enumerate(data, 3):
        is_section = label.startswith("──")
        _s(ri2, 1, label, bold=is_section, bg=bg or ("FFFFFF"), size=10 if not is_section else 11)
        if val is not None:
            _s(ri2, 2, val, bold=is_section, bg=bg or "FFFFFF", align="center", size=10)
        if note:
            _s(ri2, 3, note, bg=bg or "FFFFFF", size=9)
        ws2.row_dimensions[ri2].height = 18

    # ── Sheet 3: Confusion Matrix Visual ─────────────────────────────────────
    ws3 = wb.create_sheet("Confusion Matrix")
    ws3.sheet_view.showGridLines = False

    for col, w in [("A",4),("B",18),("C",14),("D",14),("E",4)]:
        ws3.column_dimensions[col].width = w

    ws3.merge_cells("B1:D1")
    h = ws3.cell(row=1, column=2, value="Confusion Matrix — GG≥2 Model")
    h.fill = _fill(C_HEADER_BG); h.font = _font(bold=True, color="FFFFFF", size=13)
    h.alignment = _center()

    matrix = [
        (3, 2, "",              "",         "Predicted GG≥2", "Predicted GG<2"),
        (4, 2, "Actual GG≥2",  "",         f"TP = {tp}",     f"FN = {fn}"),
        (5, 2, "Actual GG<2",  "",         f"FP = {fp}",     f"TN = {tn}"),
    ]
    colors_grid = [
        [None,       None,    C_ACCENT,   C_ACCENT],
        [C_ACCENT,   None,    C_TP,       C_FN],
        [C_ACCENT,   None,    C_FP,       C_TN],
    ]
    for i, (r, c, a, b, d, e) in enumerate(matrix):
        for j, (ci, val) in enumerate([(2,a),(3,b),(4,d),(5,e)]):
            cell = ws3.cell(row=r, column=ci, value=val)
            bg = colors_grid[i][j]
            if bg:
                cell.fill = _fill(bg)
                cell.font = _font(bold=True, color="FFFFFF" if bg==C_ACCENT else "000000", size=12)
            cell.alignment = _center()
            cell.border = _border()
            ws3.row_dimensions[r].height = 36

    # ── Sheet 4: Executive Summary for Dr. Tewari ───────────────────────────
    _build_tewari_sheet(wb, rows, counts, total, excl, evaluated,
                        tp, fp, tn, fn, precision, sensitivity, specificity,
                        reliable_total, reliable_correct,
                        unreliable_total, unreliable_correct)

    # ── Sheet 5: Plain English for Dr. Tewari ────────────────────────────────
    _build_qa_sheet(wb, rows)

    # ── Sheet 6: Dr. Tewari Q&A ──────────────────────────────────────────────
    _build_qa_sheet(wb, rows)

    # ── Sheet 6: Pathway Analysis ────────────────────────────────────────────
    _build_pathway_sheet(wb, rows)

    wb.save(out_xlsx)
    print(f"Saved → {out_xlsx}")
    print(f"  {total} patients  |  TP={tp}  FP={fp}  TN={tn}  FN={fn}")
    print(f"  Precision={precision:.2f}  Sensitivity={sensitivity:.2f}  Specificity={specificity:.2f}" if all(
        x is not None for x in [precision, sensitivity, specificity]) else "")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--in",  dest="inp", default="results/extracted.csv")
    parser.add_argument("--out", default="results/validation.xlsx")
    args = parser.parse_args()
    build_excel(args.inp, args.out)
